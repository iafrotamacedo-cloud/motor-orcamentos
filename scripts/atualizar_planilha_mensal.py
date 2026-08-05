#!/usr/bin/env python3
"""Acrescenta (dedup por ticket) uma linha na planilha mensal de orçamentos.
Colunas: Nº | Ticket | Loja | Valor total | Data.
- "Valor total" gravado como NÚMERO com formato de moeda (R$).
- "Data" = data de criação do orçamento (--data DD/MM/AAAA; padrão hoje), como DATA.

>>> BLINDAGEM CONTRA ERRO DE CONVERSÃO (texto→moeda) <<<
- O jeito MAIS SEGURO de passar o valor é `--pdf <orcamento.pdf>`: o script lê o
  `TOTAL GERAL R$ x,xx` direto do PDF, sem ninguém digitar/parsear valor à mão.
- Se usar `--valor`, aceite só formato brasileiro claro ("R$ 306,84" ou "306,84").
  Um valor SEM separador decimal e alto (ex.: "5016") é BARRADO (é o sintoma de
  vírgula perdida). Use `--inteiro` só quando o total for mesmo redondo (ex.: 600).
- NUNCA reescreve os valores das linhas já existentes (só formata) — assim um
  erro numa linha não contamina a planilha inteira.
- Guarda de sanidade: valor implausível (>= R$ 100.000 num único orçamento) é
  barrado.

Uso (acrescentar 1 linha):
  python atualizar_planilha_mensal.py --xlsx <path> --ticket 125261 --loja "MERCADÃO RUI" --pdf "orcamento.pdf" [--data 13/07/2026]
  python atualizar_planilha_mensal.py --xlsx <path> --ticket 125261 --loja "MERCADÃO RUI" --valor "R$ 306,84" [--data 13/07/2026]

Uso (REPARAR a coluna Valor total a partir dos PDFs de uma pasta):
  python atualizar_planilha_mensal.py --reparar --xlsx <path> --pdfdir "ORCAMENTOS MONTADOS/JULHO 2026"
"""
import argparse,os,sys,re,datetime,glob,subprocess
MOEDA_FMT='R$ #,##0.00'; DATA_FMT='DD/MM/YYYY'
HED=["Nº","Ticket","Loja","Valor total","Data"]
LIMITE_PLAUSIVEL=100000.0   # um orçamento acima disso é quase certamente erro

def op():
    try:
        import openpyxl;return openpyxl
    except ImportError:
        subprocess.run([sys.executable,"-m","pip","install","openpyxl","--quiet","--break-system-packages"],check=True)
        import openpyxl;return openpyxl

def parse_money_str(v):
    """Converte um texto de dinheiro BR em float. Retorna (valor, suspeito, motivo)."""
    if v is None: return (None, True, "vazio")
    if isinstance(v,(int,float)): return (round(float(v),2), False, "")
    s=re.sub(r'[^0-9,.-]','',str(v))
    if not s: return (None, True, "sem dígitos")
    tem_virg=',' in s; tem_ponto='.' in s
    if tem_virg and tem_ponto:
        s2=s.replace('.','').replace(',','.')          # 1.944,00 -> 1944.00
    elif tem_virg:
        s2=s.replace(',','.')                          # 306,84 -> 306.84
    elif tem_ponto:
        frac=s.split('.')[-1]                          # ponto só: decimal se 1-2 casas
        s2=s if (len(s.split('.'))==2 and len(frac) in (1,2)) else s.replace('.','')
    else:
        try: n=float(s)                                # SÓ DÍGITOS -> separador perdido
        except ValueError: return (None, True, "inválido")
        return (n, True, "sem separador decimal (vírgula perdida?)")
    try: return (round(float(s2),2), False, "")
    except ValueError: return (None, True, "inválido")

def valor_do_pdf(pdf):
    txt=subprocess.run(['pdftotext','-layout',pdf,'-'],capture_output=True,text=True).stdout
    m=re.search(r'TOTAL\s+GERAL\s*R\$\s*([\d.]*\d,\d{2})', txt)
    if not m: return (None, "não achei 'TOTAL GERAL R$ x,xx' no PDF")
    val,_,_=parse_money_str(m.group(1))
    return (val, "")

def parse_data(s):
    s=(s or "").strip()
    for f in ("%d/%m/%Y","%d/%m/%y","%Y-%m-%d"):
        try: return datetime.datetime.strptime(s,f).date()
        except: pass
    return None

def formatar_planilha(ws):
    from openpyxl.styles import Alignment,Border,Side
    thin=Side(style="thin",color="D9D9D9");bd=Border(left=thin,right=thin,top=thin,bottom=thin)
    for row in ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=5):
        for c in row:
            c.border=bd
            if c.row>1:
                c.alignment=Alignment(horizontal=("center" if c.column in(1,2,5) else ("right" if c.column==4 else "left")))
    for r in range(2,ws.max_row+1):
        ws.cell(row=r,column=4).number_format=MOEDA_FMT
        c5=ws.cell(row=r,column=5)
        if c5.value is not None: c5.number_format=DATA_FMT

def reparar(args, openpyxl):
    """Reescreve SÓ a coluna Valor total usando o TOTAL GERAL dos PDFs (fonte da verdade)."""
    tot={}
    for pdf in glob.glob(os.path.join(args.pdfdir,'**','*.pdf'),recursive=True):
        m=re.search(r'_(\d{5,6})\.pdf$', os.path.basename(pdf)) or re.search(r'(\d{5,6})', os.path.basename(pdf))
        if not m: continue
        v,_=valor_do_pdf(pdf)
        if v is not None: tot.setdefault(m.group(1), v)
    wb=openpyxl.load_workbook(args.xlsx);ws=wb.active
    mud=0; faltou=[]
    for r in ws.iter_rows(min_row=2):
        t=str(r[1].value).strip() if r[1].value is not None else None
        if not t: continue
        if t in tot: r[3].value=round(tot[t],2); mud+=1
        else: faltou.append(t)
    formatar_planilha(ws); wb.save(args.xlsx)
    print(f"REPARADO: {mud} valores repostos pelos PDFs | sem PDF: {len(faltou)} {faltou[:10]}")

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--xlsx",required=True)
    ap.add_argument("--reparar",action="store_true"); ap.add_argument("--pdfdir")
    ap.add_argument("--ticket"); ap.add_argument("--loja")
    ap.add_argument("--valor"); ap.add_argument("--pdf")
    ap.add_argument("--data",default=None)
    ap.add_argument("--inteiro",action="store_true",help="permite valor redondo sem centavos")
    ap.add_argument("--append",action="store_true",help="não faz dedup por ticket (permite mais de um orçamento por ticket)")
    args=ap.parse_args();openpyxl=op()
    from openpyxl.styles import Font,PatternFill,Alignment
    from openpyxl.utils import get_column_letter

    if args.reparar:
        if not args.pdfdir: sys.exit("ERRO: --reparar exige --pdfdir")
        return reparar(args, openpyxl)

    for req in ("ticket","loja"):
        if not getattr(args,req): sys.exit(f"ERRO: --{req} obrigatório")
    # ---- obtem o valor de forma segura ----
    suspeito=False; motivo=""
    if args.pdf:
        valor,err=valor_do_pdf(args.pdf)
        if valor is None: sys.exit(f"ERRO ao ler valor do PDF: {err}")
    elif args.valor is not None:
        valor,suspeito,motivo=parse_money_str(args.valor)
        if valor is None: sys.exit(f"ERRO: não entendi o valor {args.valor!r} ({motivo})")
    else:
        sys.exit("ERRO: informe --pdf <orcamento.pdf> (recomendado) ou --valor \"R$ x,xx\"")
    # ---- guardas de sanidade ----
    if suspeito and not (args.inteiro and float(valor)==int(valor)):
        sys.exit(f"BARRADO: valor {args.valor!r} parece ter PERDIDO a vírgula decimal "
                 f"({motivo}). Passe como 'R$ x,xx' (ex.: 'R$ 50,16') ou use --pdf. "
                 f"Se for MESMO redondo, repita com --inteiro.")
    if valor is None or valor<0 or valor>=LIMITE_PLAUSIVEL:
        sys.exit(f"BARRADO: valor {valor} implausível para 1 orçamento (limite R$ {LIMITE_PLAUSIVEL:,.0f}).")

    data=parse_data(args.data) or datetime.date.today()
    if os.path.exists(args.xlsx):
        wb=openpyxl.load_workbook(args.xlsx);ws=wb.active
        if (ws.cell(row=1,column=5).value or "")!="Data":
            ws.cell(row=1,column=5).value="Data"
            hc=ws.cell(row=1,column=5)
            hc.font=Font(bold=True,color="FFFFFF");hc.fill=PatternFill("solid",fgColor="1F3864")
            hc.alignment=Alignment(horizontal="center",vertical="center");ws.column_dimensions["E"].width=14
    else:
        wb=openpyxl.Workbook();ws=wb.active;ws.title="Orçamentos";ws.append(HED)
        navy=PatternFill("solid",fgColor="1F3864")
        for c in ws[1]:
            c.font=Font(bold=True,color="FFFFFF");c.fill=navy;c.alignment=Alignment(horizontal="center",vertical="center")
        ws.freeze_panes="A2"
        for i,w in enumerate([8,14,28,16,14],1): ws.column_dimensions[get_column_letter(i)].width=w

    tk=str(args.ticket).strip()
    if not args.append:
        for row in ws.iter_rows(min_row=2,values_only=True):
            if row[1] is not None and str(row[1]).strip()==tk:
                print("JA EXISTE ticket",tk,"- nada a fazer");return
    seq=ws.max_row
    ws.append([seq,tk,args.loja,round(float(valor),2),data])
    # NUNCA reprocessa valores antigos: só garante o FORMATO das colunas 4 e 5.
    formatar_planilha(ws)
    wb.save(args.xlsx)
    print(f"OK +ticket {tk} | valor R$ {valor:.2f} | data {data.strftime('%d/%m/%Y')} | total linhas: {ws.max_row-1}")

if __name__=="__main__": main()
