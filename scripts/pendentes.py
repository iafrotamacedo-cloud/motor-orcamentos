#!/usr/bin/env python3
"""
Gerencia as planilhas de NOTAS PENDENTES (as que nao viraram orcamento na 1a
passagem). Ha 1 planilha em cada pasta:
  - SEM TICKET\\NOTAS - SEM TICKET.xlsx
  - TICKET NAO ASSOCIADO\\NOTAS - TICKET NAO ASSOCIADO.xlsx

Colunas: Nº da nota | Fornecedor | Status | Ticket | Loja | RODAR
  - RODAR e uma lista suspensa SIM/NÃO (o usuario preenche).
  - Loja: o usuario preenche quando o ticket nao estara na planilha de chamados.

Subcomandos:
  add    --xlsx <path> --input <json>   (json: lista de {nota,fornecedor,status,ticket,loja})
  list   --xlsx <path>                  (imprime JSON das linhas com RODAR=SIM)
  remove --xlsx <path> --nota <n>       (remove a linha da nota)
"""
import argparse, json, os, sys

HEADERS=["Nº da nota","Fornecedor","Status","Ticket","Loja","RODAR"]
COL_NOTA=1; COL_RODAR=6

def _op():
    try:
        import openpyxl; return openpyxl
    except ImportError:
        import subprocess
        subprocess.run([sys.executable,"-m","pip","install","openpyxl","--quiet","--break-system-packages"],check=True)
        import openpyxl; return openpyxl

def _style_header(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    for c in ws[1]:
        c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1F3864")
        c.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True)
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=26
    for i,w in enumerate([12,34,20,12,28,10],1):
        ws.column_dimensions[get_column_letter(i)].width=w

def _load(openpyxl, path):
    if os.path.exists(path):
        wb=openpyxl.load_workbook(path); ws=wb.active
        if (ws.cell(row=1,column=1).value or "")!=HEADERS[0]:
            # planilha vazia/nova -> escreve cabecalho
            if ws.max_row<=1 and not ws.cell(row=1,column=1).value:
                ws.append(HEADERS); _style_header(ws)
        return wb,ws
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Pendentes"
    ws.append(HEADERS); _style_header(ws)
    return wb,ws

def _dv(ws, openpyxl):
    """(re)aplica a lista suspensa SIM/NÃO na coluna RODAR das linhas de dados."""
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    # remove DVs antigas dessa coluna e cria uma nova cobrindo ate a ultima linha
    dv=DataValidation(type="list",formula1='"SIM,NÃO"',allow_blank=True)
    ws.add_data_validation(dv)
    last=max(ws.max_row,2)
    dv.add(get_column_letter(COL_RODAR)+"2:"+get_column_letter(COL_RODAR)+str(last))

def _existentes(ws):
    m={}
    for i,row in enumerate(ws.iter_rows(min_row=2),start=2):
        n=row[0].value
        if n is not None and str(n).strip(): m[str(n).strip()]=i
    return m

def cmd_add(openpyxl,args):
    wb,ws=_load(openpyxl,args.xlsx)
    ex=_existentes(ws)
    dados=json.load(open(args.input,encoding="utf-8"))
    add=0
    for r in dados:
        nota=str(r.get("nota") or "").strip()
        if not nota or nota in ex: continue
        ws.append([nota, (r.get("fornecedor") or "").strip(), (r.get("status") or "").strip(),
                   (r.get("ticket") or "").strip(), (r.get("loja") or "").strip(), "NÃO"])
        ex[nota]=ws.max_row; add+=1
    _dv(ws,openpyxl)
    d=os.path.dirname(args.xlsx); os.makedirs(d,exist_ok=True) if d else None
    wb.save(args.xlsx)
    print("PENDENTES add: +"+str(add)+" | total: "+str(ws.max_row-1)+" | "+args.xlsx)

def cmd_list(openpyxl,args):
    if not os.path.exists(args.xlsx):
        print("[]"); return
    wb=openpyxl.load_workbook(args.xlsx); ws=wb.active
    out=[]
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not row or not (row[0] and str(row[0]).strip()): continue
        rodar=str((row[COL_RODAR-1] or "")).strip().upper()
        if rodar=="SIM":
            out.append({"nota":str(row[0]).strip(),"fornecedor":row[1],"status":row[2],
                        "ticket":(str(row[3]).strip() if row[3] is not None else ""),
                        "loja":(str(row[4]).strip() if row[4] is not None else "")})
    print(json.dumps(out,ensure_ascii=False))

def cmd_remove(openpyxl,args):
    if not os.path.exists(args.xlsx): return
    wb=openpyxl.load_workbook(args.xlsx); ws=wb.active
    ex=_existentes(ws)
    n=str(args.nota).strip()
    if n in ex:
        ws.delete_rows(ex[n],1); _dv(ws,openpyxl); wb.save(args.xlsx)
        print("PENDENTES remove: "+n+" removido | total: "+str(ws.max_row-1))
    else:
        print("PENDENTES remove: "+n+" nao encontrado")

def main():
    ap=argparse.ArgumentParser()
    sub=ap.add_subparsers(dest="cmd",required=True)
    a=sub.add_parser("add"); a.add_argument("--xlsx",required=True); a.add_argument("--input",required=True)
    l=sub.add_parser("list"); l.add_argument("--xlsx",required=True)
    r=sub.add_parser("remove"); r.add_argument("--xlsx",required=True); r.add_argument("--nota",required=True)
    args=ap.parse_args()
    openpyxl=_op()
    {"add":cmd_add,"list":cmd_list,"remove":cmd_remove}[args.cmd](openpyxl,args)

if __name__=="__main__":
    main()
